from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_TEST_JOBS_INPUT = PROJECT_ROOT / "data" / "test" / "job_descriptions_test.jsonl"
DEFAULT_SSOC_XLSX = PROJECT_ROOT / "data" / "ssoc2020.xlsx"
DEFAULT_OUTPUT_JSONL = PROJECT_ROOT / "data" / "test" / "job_ssoc345_with_skills_from_original.jsonl"
DEFAULT_OUTPUT_JSON = PROJECT_ROOT / "data" / "test" / "job_ssoc345_with_skills_from_original.json"


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_ssoc_title_lookup(xlsx_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = normalize_text(row[0] if len(row) > 0 else "")
        if not (code.isdigit() and len(code) in {3, 4, 5}):
            continue

        lookup[code] = normalize_text(row[1] if len(row) > 1 else "")
    return lookup


def parse_ssoc_code(raw) -> tuple[str | None, str | None, str | None]:
    text = normalize_text(raw)
    if not text:
        return None, None, None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 5:
        return None, None, None
    ssoc5 = digits[:5]
    ssoc4 = digits[:4]
    ssoc3 = digits[:3]
    return ssoc5, ssoc4, ssoc3


def extract_skill_names(row: dict) -> list[str]:
    names = []
    seen = set()
    for key in ("skills", "all_relevant_skills", "hard_skills", "soft_skills"):
        for item in row.get(key) or []:
            skill = normalize_text(item)
            if not skill:
                continue
            skill_key = skill.lower()
            if skill_key in seen:
                continue
            seen.add(skill_key)
            names.append(skill)
    return names


def parse_binary_flag(value) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value != 0)

    text = normalize_text(value).lower()
    if text in {"1", "true", "yes", "y"}:
        return 1
    return 0


def load_cleaned_rows(input_path: Path) -> list[dict]:
    suffix = input_path.suffix.lower()
    if suffix == ".jsonl":
        rows = []
        with input_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        return rows
    if suffix == ".json":
        payload = json.loads(input_path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(
                f"Expected a JSON array in {input_path}, got {type(payload).__name__}."
            )
        return payload

    raise ValueError(
        f"Unsupported input format: {input_path}. Expected .jsonl or .json."
    )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Extract SSOC hierarchy from downstream job test rows and map to SSOC 2020 Excel."
        )
    )
    parser.add_argument("--jobs-input", type=Path, default=DEFAULT_TEST_JOBS_INPUT)
    parser.add_argument("--ssoc-xlsx", type=Path, default=DEFAULT_SSOC_XLSX)
    parser.add_argument("--output-jsonl", type=Path, default=DEFAULT_OUTPUT_JSONL)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--include-non-fresh", action="store_true")
    args = parser.parse_args()

    if not args.jobs_input.exists():
        raise FileNotFoundError(f"Job test input not found: {args.jobs_input}")
    if not args.ssoc_xlsx.exists():
        raise FileNotFoundError(f"SSOC Excel not found: {args.ssoc_xlsx}")

    ssoc_title_lookup = load_ssoc_title_lookup(args.ssoc_xlsx)
    if not ssoc_title_lookup:
        raise RuntimeError("No SSOC 3/4/5-digit rows found in the Excel file.")

    cleaned_rows = load_cleaned_rows(args.jobs_input)

    fresh_only = not args.include_non_fresh
    mapped_rows: list[dict] = []
    skipped_no_ssoc = 0
    skipped_no_match = 0
    skipped_non_fresh = 0
    ssoc3_total_counts: Counter[str] = Counter()
    ssoc3_good_counts: Counter[str] = Counter()

    total = len(cleaned_rows)
    for i, row in enumerate(cleaned_rows, start=1):
        has_fresh_flag = "is_freshgrad" in row
        if fresh_only and has_fresh_flag and not row.get("is_freshgrad"):
            skipped_non_fresh += 1
            continue

        ssoc5, ssoc4, ssoc3 = parse_ssoc_code(row.get("ssoc_code") or row.get("ssocCode"))
        if not ssoc5:
            skipped_no_ssoc += 1
            continue

        title5 = ssoc_title_lookup.get(ssoc5, "")
        title4 = ssoc_title_lookup.get(ssoc4, "")
        title3 = ssoc_title_lookup.get(ssoc3, "")
        if not title3:
            skipped_no_match += 1
            continue

        is_good_job = parse_binary_flag(row.get("is_good_job"))
        ssoc3_total_counts[ssoc3] += 1
        ssoc3_good_counts[ssoc3] += is_good_job

        mapped_rows.append(
            {
                "job_post_id": row.get("id") or row.get("uuid"),
                "ssoc_5d_code": ssoc5,
                "ssoc_5d_title": title5,
                "ssoc_4d_code": ssoc4,
                "ssoc_4d_title": title4,
                "ssoc_3d_code": ssoc3,
                "ssoc_3d_title": title3,
                "is_good_job": is_good_job,
                "skills": extract_skill_names(row),
            }
        )

        if i % 5000 == 0 or i == total:
            print(f"Scanned {i}/{total} rows...", flush=True)

    for mapped_row in mapped_rows:
        ssoc3 = mapped_row["ssoc_3d_code"]
        total_jobs_in_group = ssoc3_total_counts.get(ssoc3, 0)
        good_jobs_in_group = ssoc3_good_counts.get(ssoc3, 0)
        mapped_row["ssoc_3d_total_jobs"] = total_jobs_in_group
        mapped_row["ssoc_3d_good_jobs"] = good_jobs_in_group
        good_job_pct = (
            (good_jobs_in_group / total_jobs_in_group)
            if total_jobs_in_group
            else 0.0
        )
        mapped_row["ssoc_3d_good_job_pct"] = round(good_job_pct, 4)

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as f:
        for mapped_row in mapped_rows:
            f.write(json.dumps(mapped_row, ensure_ascii=False) + "\n")

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(mapped_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Total rows scanned: {total}")
    print(f"Mapped rows: {len(mapped_rows)}")
    if fresh_only and any("is_freshgrad" in row for row in cleaned_rows):
        print(f"Skipped (non-fresh rows): {skipped_non_fresh}")
    print(f"Skipped (missing ssocCode): {skipped_no_ssoc}")
    print(f"Skipped (3-digit hierarchy not in SSOC file): {skipped_no_match}")
    print(f"Saved JSONL: {args.output_jsonl}")
    print(f"Saved JSON: {args.output_json}")


if __name__ == "__main__":
    main()
