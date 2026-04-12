import argparse
import ast
import json
import math
import re
import heapq
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DEFAULT_MODULE_SKILLS = PROJECT_ROOT / "data" / "reupdated_skills.xlsx"
DEFAULT_JOB_SKILLS = PROJECT_ROOT / "data" / "processed" / "jobs_cleaned.json"

DEFAULT_NUS_CLEANED = PROJECT_ROOT / "data" / "processed" / "nus_cleaned.json"
DEFAULT_NTU_CLEANED = PROJECT_ROOT / "data" / "processed" / "ntu_cleaned.json"
DEFAULT_SUTD_CLEANED = PROJECT_ROOT / "data" / "processed" / "sutd_cleaned.json"

DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "test"
DEFAULT_MODULE_ALIGNMENT_OUT = DEFAULT_OUTPUT_DIR / "module_job_alignment.json"
DEFAULT_GROUP_ALIGNMENT_OUT = DEFAULT_OUTPUT_DIR / "module_group_job_alignment.json"
DEFAULT_PROGRAMME_ALIGNMENT_OUT = DEFAULT_OUTPUT_DIR / "programme_job_alignment.json"
DEFAULT_SEMANTIC_MAP_OUT = DEFAULT_OUTPUT_DIR / "skill_semantic_mapping.json"
DEFAULT_MODULE_KEYBERT_SKILLS_JSON = DEFAULT_OUTPUT_DIR / "module_descriptions_test_with_skills_keybert.json"
DEFAULT_MODULE_TRANSFORMER_SKILLS_JSON = DEFAULT_OUTPUT_DIR / "module_descriptions_test_with_skills.json"


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def parse_list_like(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(x) for x in value]
    if isinstance(value, tuple):
        return [str(x) for x in value]
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return []
        try:
            obj = ast.literal_eval(s)
            if isinstance(obj, (list, tuple)):
                return [str(x) for x in obj]
        except Exception:
            pass
        return [s]
    return [str(value)]


def normalize_label(value: str | None) -> str:
    if not value:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def maybe_map_skill(skill: str, skill_mapper: dict[str, str] | None) -> str:
    if not skill_mapper:
        return skill
    return skill_mapper.get(skill, skill)


def normalize_skill(skill: str | None) -> str:
    if not skill:
        return ""
    s = str(skill).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"^[^a-z0-9+#/]+|[^a-z0-9+#/]+$", "", s)
    return s


def normalize_counter(counter: Counter) -> dict[str, float]:
    total = sum(counter.values())
    if total <= 0:
        return {}
    return {k: v / total for k, v in counter.items()}


def cosine_similarity(a: dict[str, float], b: dict[str, float]) -> float:
    if not a or not b:
        return 0.0
    common = set(a).intersection(b)
    dot = sum(a[k] * b[k] for k in common)
    na = math.sqrt(sum(v * v for v in a.values()))
    nb = math.sqrt(sum(v * v for v in b.values()))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)


def weighted_jaccard(a: dict[str, float], b: dict[str, float]) -> float:
    keys = set(a).union(b)
    if not keys:
        return 0.0
    num = sum(min(a.get(k, 0.0), b.get(k, 0.0)) for k in keys)
    den = sum(max(a.get(k, 0.0), b.get(k, 0.0)) for k in keys)
    if den == 0:
        return 0.0
    return num / den


def topk_coverage(module_counter: Counter, job_counter: Counter, top_k: int) -> float:
    top_job = [s for s, _ in job_counter.most_common(top_k)]
    if not top_job:
        return 0.0
    module_skills = set(module_counter)
    covered = sum(1 for s in top_job if s in module_skills)
    return covered / len(top_job)


def gap_score(module_counter: Counter, job_counter: Counter, top_k: int) -> float:
    top_job = job_counter.most_common(top_k)
    if not top_job:
        return 0.0
    total_weight = sum(w for _, w in top_job)
    if total_weight <= 0:
        return 0.0
    module_skills = set(module_counter)
    missing_weight = sum(w for s, w in top_job if s not in module_skills)
    return missing_weight / total_weight


def module_counter_from_skills(skills: list[str]) -> Counter:
    counter = Counter()
    for s in skills:
        ns = normalize_skill(s)
        ns = maybe_map_skill(ns, None)
        if ns:
            counter[ns] += 1
    return counter


def module_counter_from_skills_with_mapper(skills: list[str], skill_mapper: dict[str, str] | None) -> Counter:
    counter = Counter()
    for s in skills:
        ns = normalize_skill(s)
        ns = maybe_map_skill(ns, skill_mapper)
        if ns:
            counter[ns] += 1
    return counter


def collect_job_skill_counter(job_rows: list[dict], freshgrad_only: bool, job_skill_field: str) -> Counter:
    c = Counter()
    for row in job_rows:
        if freshgrad_only and row.get("is_freshgrad") is not True:
            continue
        raw_skills = row.get(job_skill_field)
        if raw_skills is None:
            raw_skills = row.get("skills") or row.get("all_relevant_skills") or []
        for s in raw_skills or []:
            ns = normalize_skill(s)
            if ns:
                c[ns] += 1
    return c


def collect_module_unique_skills(module_rows: list[dict]) -> list[str]:
    s = set()
    for row in module_rows:
        for skill in row.get("skills") or []:
            ns = normalize_skill(skill)
            if ns:
                s.add(ns)
    return sorted(s)


def build_semantic_skill_mapper(
    module_skills: list[str],
    canonical_skill_counter: Counter,
    model_name: str,
    threshold: float,
    max_canonical: int,
    batch_size: int,
    local_files_only: bool,
) -> tuple[dict[str, str], list[dict]]:
    canonical_skills = [s for s, _ in canonical_skill_counter.most_common(max_canonical)]
    canonical_set = set(canonical_skills)

    mapper: dict[str, str] = {}
    audit: list[dict] = []

    # First pass: exact matches
    pending = []
    for s in module_skills:
        if s in canonical_set:
            mapper[s] = s
            audit.append(
                {
                    "raw_skill": s,
                    "mapped_skill": s,
                    "match_type": "exact",
                    "similarity": 1.0,
                }
            )
        else:
            pending.append(s)

    if not pending or not canonical_skills:
        for s in pending:
            mapper[s] = s
            audit.append(
                {
                    "raw_skill": s,
                    "mapped_skill": s,
                    "match_type": "unmapped",
                    "similarity": 0.0,
                }
            )
        return mapper, audit

    try:
        from sentence_transformers import SentenceTransformer
    except Exception as e:
        # Fallback to identity mapping if semantic model can't be loaded.
        for s in pending:
            mapper[s] = s
            audit.append(
                {
                    "raw_skill": s,
                    "mapped_skill": s,
                    "match_type": "unmapped_no_model",
                    "similarity": 0.0,
                    "note": str(e),
                }
            )
        return mapper, audit

    try:
        model = SentenceTransformer(model_name, local_files_only=local_files_only)
    except Exception as e:
        for s in pending:
            mapper[s] = s
            audit.append(
                {
                    "raw_skill": s,
                    "mapped_skill": s,
                    "match_type": "unmapped_model_load_error",
                    "similarity": 0.0,
                    "note": str(e),
                }
            )
        return mapper, audit
    canon_emb = model.encode(
        canonical_skills,
        batch_size=batch_size,
        convert_to_numpy=True,
        normalize_embeddings=True,
        show_progress_bar=False,
    )

    for i in range(0, len(pending), batch_size):
        batch = pending[i : i + batch_size]
        raw_emb = model.encode(
            batch,
            batch_size=batch_size,
            convert_to_numpy=True,
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        sims = raw_emb @ canon_emb.T
        best_idx = np.argmax(sims, axis=1)
        best_sim = np.max(sims, axis=1)

        for raw_skill, idx, sim in zip(batch, best_idx.tolist(), best_sim.tolist()):
            mapped = canonical_skills[idx] if sim >= threshold else raw_skill
            match_type = "semantic" if sim >= threshold else "unmapped_below_threshold"
            mapper[raw_skill] = mapped
            audit.append(
                {
                    "raw_skill": raw_skill,
                    "mapped_skill": mapped,
                    "match_type": match_type,
                    "similarity": round(float(sim), 6),
                }
            )

    return mapper, audit


def load_module_rows(path: Path, module_skill_fields: list[str]) -> list[dict]:
    if path.suffix.lower() in {".json", ".jsonl"}:
        rows = load_json(path)
        if not isinstance(rows, list):
            raise ValueError(f"Expected JSON array in {path}")
        return rows

    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter)
        headers = [str(h).strip() if h is not None else "" for h in header]
        idx = {name: i for i, name in enumerate(headers)}

        out = []
        for r in rows_iter:
            code = r[idx["code"]] if "code" in idx else None
            university = r[idx["university"]] if "university" in idx else None
            title = r[idx["title"]] if "title" in idx else None
            department = r[idx["department"]] if "department" in idx else None
            if not code or not university:
                continue

            skills = []
            for field in module_skill_fields:
                if field not in idx:
                    continue
                skills.extend(parse_list_like(r[idx[field]]))

            out.append(
                {
                    "id": f"{str(university).strip()}::{str(code).strip()}",
                    "source": str(university).strip(),
                    "title": title,
                    "department": department,
                    "skills": skills,
                }
            )
        return out

    raise ValueError(f"Unsupported module skills file type: {path.suffix}")


def load_module_skill_json_map(path: Path) -> dict[str, list[str]]:
    rows = load_json(path)
    if not isinstance(rows, list):
        raise ValueError(f"Expected JSON array in {path}")

    out: dict[str, list[str]] = {}
    for row in rows:
        module_id = row.get("id")
        if not module_id:
            continue
        skills = parse_list_like(row.get("skills"))
        out[str(module_id)] = skills
    return out


def merge_module_skill_sources(
    module_rows: list[dict],
    extra_skill_maps: list[dict[str, list[str]]],
    extra_min_support: int,
) -> tuple[list[dict], dict[str, int]]:
    if not extra_skill_maps:
        def _unique_count(skills: list[str]) -> int:
            return len({normalize_skill(s) for s in skills if normalize_skill(s)})

        return module_rows, {
            "modules_with_extra_source_hit": 0,
            "base_unique_skill_total": sum(_unique_count(parse_list_like(r.get("skills"))) for r in module_rows),
            "merged_unique_skill_total": sum(_unique_count(parse_list_like(r.get("skills"))) for r in module_rows),
        }

    merged_rows: list[dict] = []
    modules_with_extra_source_hit = 0
    base_unique_skill_total = 0
    merged_unique_skill_total = 0

    for row in module_rows:
        module_id = row.get("id")
        base_skills = parse_list_like(row.get("skills"))

        base_seen = set()
        ordered_base = []
        for s in base_skills:
            ns = normalize_skill(s)
            if not ns or ns in base_seen:
                continue
            base_seen.add(ns)
            ordered_base.append(s)
        base_unique_skill_total += len(base_seen)

        votes: Counter = Counter()
        any_extra_hit = False
        for skill_map in extra_skill_maps:
            extra_skills = skill_map.get(module_id, []) if module_id else []
            normalized_unique = set()
            for s in extra_skills:
                ns = normalize_skill(s)
                if ns:
                    normalized_unique.add(ns)
            if normalized_unique:
                any_extra_hit = True
            for ns in normalized_unique:
                votes[ns] += 1

        if any_extra_hit:
            modules_with_extra_source_hit += 1

        merged_skills = list(ordered_base)
        merged_seen = set(normalize_skill(s) for s in ordered_base if normalize_skill(s))

        # Consensus add-on: skills from auxiliary extractors that appear in at least N sources.
        for ns, cnt in votes.most_common():
            if cnt < max(1, extra_min_support):
                continue
            if ns in merged_seen:
                continue
            merged_seen.add(ns)
            merged_skills.append(ns)

        merged_unique_skill_total += len(merged_seen)
        merged_rows.append({**row, "skills": merged_skills})

    stats = {
        "modules_with_extra_source_hit": modules_with_extra_source_hit,
        "base_unique_skill_total": base_unique_skill_total,
        "merged_unique_skill_total": merged_unique_skill_total,
    }
    return merged_rows, stats


def build_job_profiles(
    job_rows: list[dict],
    freshgrad_only: bool,
    job_skill_field: str,
    skill_mapper: dict[str, str] | None = None,
) -> dict[str, dict]:
    counters: dict[str, Counter] = {}
    titles: dict[str, str] = {}
    for row in job_rows:
        if freshgrad_only and row.get("is_freshgrad") is not True:
            continue

        # Prefer explicit SSOC-3 if present; otherwise derive from ssoc_code.
        code = row.get("ssoc_3d_code")
        if not code:
            ssoc_code = str(row.get("ssoc_code") or "").strip()
            digits = "".join(ch for ch in ssoc_code if ch.isdigit())
            code = digits[:3] if len(digits) >= 3 else None
        title = row.get("ssoc_3d_title") or (f"SSOC-{code}" if code else "")
        if not code:
            continue

        if code not in counters:
            counters[code] = Counter()
            titles[code] = title

        raw_skills = row.get(job_skill_field)
        if raw_skills is None:
            raw_skills = row.get("skills") or row.get("all_relevant_skills") or []

        for s in set(maybe_map_skill(normalize_skill(x), skill_mapper) for x in (raw_skills or [])):
            if s:
                counters[code][s] += 1

    profiles = {}
    for code, c in counters.items():
        profiles[code] = {
            "ssoc_3d_code": code,
            "ssoc_3d_title": titles.get(code, ""),
            "skill_counter": c,
            "skill_weights": normalize_counter(c),
        }
    return profiles


def build_job_listing_profiles(
    job_rows: list[dict],
    freshgrad_only: bool,
    job_skill_field: str,
    skill_mapper: dict[str, str] | None = None,
) -> dict[str, list[dict]]:
    by_group: dict[str, list[dict]] = defaultdict(list)

    for row in job_rows:
        if freshgrad_only and row.get("is_freshgrad") is not True:
            continue

        code = row.get("ssoc_3d_code")
        if not code:
            ssoc_code = str(row.get("ssoc_code") or "").strip()
            digits = "".join(ch for ch in ssoc_code if ch.isdigit())
            code = digits[:3] if len(digits) >= 3 else None
        if not code:
            continue

        raw_skills = row.get(job_skill_field)
        if raw_skills is None:
            raw_skills = row.get("skills") or row.get("all_relevant_skills") or []
        counter = module_counter_from_skills_with_mapper(raw_skills or [], skill_mapper)
        if not counter:
            continue

        job_id = row.get("id") or row.get("job_post_id")
        title = row.get("title")
        if not job_id:
            continue

        by_group[code].append(
            {
                "id": str(job_id),
                "title": title,
                "ssoc_3d_code": code,
                "skill_counter": counter,
                "skill_weights": normalize_counter(counter),
            }
        )
    return by_group


def build_module_meta_lookup(nus_path: Path, ntu_path: Path, sutd_path: Path) -> dict[str, dict]:
    lookup = {}

    if nus_path.exists():
        for r in load_json(nus_path):
            code = r.get("moduleCode")
            if code:
                lookup[f"NUS::{code}"] = {
                    "department": r.get("department"),
                    "title": r.get("title"),
                    "programme_tags": [normalize_label(r.get("department"))] if normalize_label(r.get("department")) else [],
                }

    if ntu_path.exists():
        for r in load_json(ntu_path):
            code = r.get("code")
            if code:
                lookup[f"NTU::{code}"] = {
                    "department": r.get("department"),
                    "title": r.get("title"),
                    "programme_tags": [normalize_label(r.get("department"))] if normalize_label(r.get("department")) else [],
                }

    if sutd_path.exists():
        for r in load_json(sutd_path):
            code = r.get("code")
            if code:
                dept = r.get("department")
                if isinstance(dept, list):
                    dept = dept[0] if dept else None
                tags = []
                tags.extend(parse_list_like(r.get("affiliations")))
                tags.append(r.get("course_type"))
                tags.append(dept)
                normalized_tags = []
                seen = set()
                for t in tags:
                    nt = normalize_label(t)
                    key = nt.lower()
                    if not nt or key in seen:
                        continue
                    seen.add(key)
                    normalized_tags.append(nt)

                lookup[f"SUTD::{code}"] = {
                    "department": dept,
                    "title": r.get("title"),
                    "programme_tags": normalized_tags,
                }
    return lookup


def resolve_programmes(row: dict, meta: dict) -> list[str]:
    candidates = []
    for key in ("programme", "program", "major", "track", "school"):
        candidates.extend(parse_list_like(row.get(key)))
    candidates.extend(parse_list_like(meta.get("programme_tags")))
    candidates.append(row.get("department"))
    candidates.append(meta.get("department"))

    out = []
    seen = set()
    for c in candidates:
        nc = normalize_label(c)
        key = nc.lower()
        if not nc or key in seen:
            continue
        seen.add(key)
        out.append(nc)
    return out


def score_entity_against_jobs(
    entity_counter: Counter,
    job_profiles: dict[str, dict],
    top_k_job_skills: int,
    top_n_matches: int,
) -> list[dict]:
    entity_weights = normalize_counter(entity_counter)
    matches = []

    for code, jp in job_profiles.items():
        j_counter = jp["skill_counter"]
        j_weights = jp["skill_weights"]

        cov = topk_coverage(entity_counter, j_counter, top_k_job_skills)
        wj = weighted_jaccard(entity_weights, j_weights)
        cos = cosine_similarity(entity_weights, j_weights)
        gap = gap_score(entity_counter, j_counter, top_k_job_skills)
        alignment = 0.35 * cov + 0.35 * wj + 0.30 * cos - 0.20 * gap

        top_job = [s for s, _ in j_counter.most_common(top_k_job_skills)]
        entity_skills = set(entity_counter)
        missing = [s for s in top_job if s not in entity_skills][:10]

        matches.append(
            {
                "ssoc_3d_code": code,
                "ssoc_3d_title": jp["ssoc_3d_title"],
                "coverage_top_k": round(cov, 4),
                "weighted_jaccard": round(wj, 4),
                "cosine_similarity": round(cos, 4),
                "gap_score": round(gap, 4),
                "alignment_score": round(alignment, 4),
                "top_missing_skills": missing,
            }
        )

    matches.sort(key=lambda x: x["alignment_score"], reverse=True)
    return matches[:top_n_matches]


def score_entity_to_job_listing(
    entity_counter: Counter,
    entity_weights: dict[str, float],
    job_counter: Counter,
    job_weights: dict[str, float],
    top_k_job_skills: int,
) -> tuple[float, float, float, float, float, list[str]]:
    cov = topk_coverage(entity_counter, job_counter, top_k_job_skills)
    wj = weighted_jaccard(entity_weights, job_weights)
    cos = cosine_similarity(entity_weights, job_weights)
    gap = gap_score(entity_counter, job_counter, top_k_job_skills)
    alignment = 0.35 * cov + 0.35 * wj + 0.30 * cos - 0.20 * gap

    top_job = [s for s, _ in job_counter.most_common(top_k_job_skills)]
    entity_skills = set(entity_counter)
    missing = [s for s in top_job if s not in entity_skills][:10]
    return cov, wj, cos, gap, alignment, missing


def top_job_listings_for_entity(
    entity_counter: Counter,
    job_group_matches: list[dict],
    job_listings_by_group: dict[str, list[dict]],
    candidate_groups: int,
    top_job_listings: int,
    top_k_job_skills: int,
) -> list[dict]:
    entity_weights = normalize_counter(entity_counter)
    candidate_codes = [m["ssoc_3d_code"] for m in job_group_matches[:candidate_groups]]
    candidates = []
    for code in candidate_codes:
        candidates.extend(job_listings_by_group.get(code, []))

    if not candidates:
        return []

    scored = []
    for job in candidates:
        cov, wj, cos, gap, alignment, missing = score_entity_to_job_listing(
            entity_counter=entity_counter,
            entity_weights=entity_weights,
            job_counter=job["skill_counter"],
            job_weights=job["skill_weights"],
            top_k_job_skills=top_k_job_skills,
        )
        scored.append(
            {
                "job_id": job["id"],
                "title": job["title"],
                "ssoc_3d_code": job["ssoc_3d_code"],
                "coverage_top_k": round(cov, 4),
                "weighted_jaccard": round(wj, 4),
                "cosine_similarity": round(cos, 4),
                "gap_score": round(gap, 4),
                "alignment_score": round(alignment, 4),
                "top_missing_skills": missing,
            }
        )

    return heapq.nlargest(top_job_listings, scored, key=lambda x: x["alignment_score"])


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Align module skills directly with job skill profiles (SSOC-3), "
            "without simulating student profiles."
        )
    )
    parser.add_argument("--module-skills", type=Path, default=DEFAULT_MODULE_SKILLS)
    parser.add_argument(
        "--module-keybert-skills-json",
        type=Path,
        default=None,
        help=(
            "Optional JSON with module-level extracted skills (id, skills), "
            "e.g. output from extract_module_skills_keybert.py."
        ),
    )
    parser.add_argument(
        "--module-transformer-skills-json",
        type=Path,
        default=None,
        help=(
            "Optional JSON with module-level extracted skills (id, skills), "
            "e.g. output from extract_module_skills_transformer.py."
        ),
    )
    parser.add_argument(
        "--module-extra-skills-min-support",
        type=int,
        default=1,
        help=(
            "Minimum number of auxiliary sources (KeyBERT/Transformer/etc.) that must agree "
            "before adding an extra skill not already present in the base module skills."
        ),
    )
    parser.add_argument("--job-skills", type=Path, default=DEFAULT_JOB_SKILLS)
    parser.add_argument("--nus-cleaned", type=Path, default=DEFAULT_NUS_CLEANED)
    parser.add_argument("--ntu-cleaned", type=Path, default=DEFAULT_NTU_CLEANED)
    parser.add_argument("--sutd-cleaned", type=Path, default=DEFAULT_SUTD_CLEANED)
    parser.add_argument(
        "--module-skill-fields",
        type=str,
        default="skills,hard_skills_v2,soft_skills_v2,skills_embedding_v2",
        help="Comma-separated skill columns for Excel module input.",
    )
    parser.add_argument(
        "--semantic-map-skills",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Map module/job skills to nearest canonical job skill using embeddings.",
    )
    parser.add_argument("--semantic-model", type=str, default="sentence-transformers/all-MiniLM-L6-v2")
    parser.add_argument(
        "--semantic-local-only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Load semantic model from local cache only (recommended in restricted network envs).",
    )
    parser.add_argument("--semantic-threshold", type=float, default=0.62)
    parser.add_argument("--semantic-max-canonical", type=int, default=8000)
    parser.add_argument("--semantic-batch-size", type=int, default=512)
    parser.add_argument("--semantic-map-out", type=Path, default=DEFAULT_SEMANTIC_MAP_OUT)
    parser.add_argument("--top-k-job-skills", type=int, default=20)
    parser.add_argument("--top-n-matches", type=int, default=10)
    parser.add_argument("--job-skill-field", type=str, default="all_relevant_skills")
    parser.add_argument(
        "--freshgrad-only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="When true, only include jobs with is_freshgrad=True if the field exists.",
    )
    parser.add_argument("--module-alignment-out", type=Path, default=DEFAULT_MODULE_ALIGNMENT_OUT)
    parser.add_argument("--group-alignment-out", type=Path, default=DEFAULT_GROUP_ALIGNMENT_OUT)
    parser.add_argument("--programme-alignment-out", type=Path, default=DEFAULT_PROGRAMME_ALIGNMENT_OUT)
    parser.add_argument("--top-job-listings", type=int, default=5)
    parser.add_argument(
        "--candidate-groups-for-listings",
        type=int,
        default=3,
        help="Rerank listings only inside top-N job groups per module.",
    )
    args = parser.parse_args()

    if not args.module_skills.exists():
        raise FileNotFoundError(f"Module skills file not found: {args.module_skills}")
    if not args.job_skills.exists():
        raise FileNotFoundError(f"Job skills file not found: {args.job_skills}")

    module_skill_fields = [x.strip() for x in args.module_skill_fields.split(",") if x.strip()]
    module_rows = load_module_rows(args.module_skills, module_skill_fields=module_skill_fields)
    extra_skill_maps = []
    if args.module_keybert_skills_json:
        if not args.module_keybert_skills_json.exists():
            raise FileNotFoundError(f"Module KeyBERT skills file not found: {args.module_keybert_skills_json}")
        extra_skill_maps.append(load_module_skill_json_map(args.module_keybert_skills_json))
    if args.module_transformer_skills_json:
        if not args.module_transformer_skills_json.exists():
            raise FileNotFoundError(
                f"Module transformer skills file not found: {args.module_transformer_skills_json}"
            )
        extra_skill_maps.append(load_module_skill_json_map(args.module_transformer_skills_json))

    module_rows, merge_stats = merge_module_skill_sources(
        module_rows=module_rows,
        extra_skill_maps=extra_skill_maps,
        extra_min_support=args.module_extra_skills_min_support,
    )
    module_source_type = "excel" if args.module_skills.suffix.lower() == ".xlsx" else "json"
    job_rows = load_json(args.job_skills)
    skill_mapper = None
    semantic_audit = []

    if args.semantic_map_skills:
        module_unique_skills = collect_module_unique_skills(module_rows)
        canonical_counter = collect_job_skill_counter(
            job_rows=job_rows,
            freshgrad_only=args.freshgrad_only,
            job_skill_field=args.job_skill_field,
        )
        skill_mapper, semantic_audit = build_semantic_skill_mapper(
            module_skills=module_unique_skills,
            canonical_skill_counter=canonical_counter,
            model_name=args.semantic_model,
            threshold=args.semantic_threshold,
            max_canonical=args.semantic_max_canonical,
            batch_size=args.semantic_batch_size,
            local_files_only=args.semantic_local_only,
        )

    job_profiles = build_job_profiles(
        job_rows=job_rows,
        freshgrad_only=args.freshgrad_only,
        job_skill_field=args.job_skill_field,
        skill_mapper=skill_mapper,
    )
    job_listings_by_group = build_job_listing_profiles(
        job_rows=job_rows,
        freshgrad_only=args.freshgrad_only,
        job_skill_field=args.job_skill_field,
        skill_mapper=skill_mapper,
    )
    module_meta = build_module_meta_lookup(args.nus_cleaned, args.ntu_cleaned, args.sutd_cleaned)

    module_alignment = []
    grouped_skill_counters: dict[str, Counter] = defaultdict(Counter)
    grouped_meta: dict[str, dict] = {}
    programme_skill_counters: dict[str, Counter] = defaultdict(Counter)
    programme_meta: dict[str, dict] = {}

    for row in module_rows:
        module_id = row.get("id")
        if not module_id:
            continue

        skills = row.get("skills") or []
        counter = module_counter_from_skills_with_mapper(skills, skill_mapper)
        if not counter:
            continue

        source = row.get("source")
        meta = module_meta.get(module_id, {})
        department = row.get("department") or meta.get("department")
        title = row.get("title") or meta.get("title")

        top_matches = score_entity_against_jobs(
            entity_counter=counter,
            job_profiles=job_profiles,
            top_k_job_skills=args.top_k_job_skills,
            top_n_matches=args.top_n_matches,
        )
        top_listings = top_job_listings_for_entity(
            entity_counter=counter,
            job_group_matches=top_matches,
            job_listings_by_group=job_listings_by_group,
            candidate_groups=args.candidate_groups_for_listings,
            top_job_listings=args.top_job_listings,
            top_k_job_skills=args.top_k_job_skills,
        )
        module_alignment.append(
            {
                "module_id": module_id,
                "source": source,
                "department": department,
                "title": title,
                "num_unique_skills": len(counter),
                "top_skills": [s for s, _ in counter.most_common(20)],
                "top_matches": top_matches,
                "top_job_listings": top_listings,
            }
        )

        group_key = f"{source}::{department}" if department else f"{source}::UNKNOWN"
        grouped_skill_counters[group_key].update(counter)
        grouped_meta[group_key] = {"source": source, "department": department}

        programmes = resolve_programmes(row=row, meta=meta)
        if not programmes:
            programmes = [department] if department else ["UNKNOWN"]
        for programme in programmes:
            programme_key = f"{source}::{programme}"
            programme_skill_counters[programme_key].update(counter)
            programme_meta[programme_key] = {"source": source, "programme": programme}

    group_alignment = []
    for group_key, counter in grouped_skill_counters.items():
        gm = grouped_meta[group_key]
        top_matches = score_entity_against_jobs(
            entity_counter=counter,
            job_profiles=job_profiles,
            top_k_job_skills=args.top_k_job_skills,
            top_n_matches=args.top_n_matches,
        )
        group_alignment.append(
            {
                "group_id": group_key,
                "source": gm["source"],
                "department": gm["department"],
                "num_unique_skills": len(counter),
                "top_skills": [s for s, _ in counter.most_common(30)],
                "top_matches": top_matches,
            }
        )

    programme_alignment = []
    for programme_key, counter in programme_skill_counters.items():
        pm = programme_meta[programme_key]
        top_matches = score_entity_against_jobs(
            entity_counter=counter,
            job_profiles=job_profiles,
            top_k_job_skills=args.top_k_job_skills,
            top_n_matches=args.top_n_matches,
        )
        programme_alignment.append(
            {
                "programme_id": programme_key,
                "source": pm["source"],
                "programme": pm["programme"],
                "num_unique_skills": len(counter),
                "top_skills": [s for s, _ in counter.most_common(30)],
                "top_matches": top_matches,
            }
        )

    args.module_alignment_out.parent.mkdir(parents=True, exist_ok=True)
    args.group_alignment_out.parent.mkdir(parents=True, exist_ok=True)
    args.programme_alignment_out.parent.mkdir(parents=True, exist_ok=True)
    args.module_alignment_out.write_text(
        json.dumps(module_alignment, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    args.group_alignment_out.write_text(
        json.dumps(group_alignment, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    args.programme_alignment_out.write_text(
        json.dumps(programme_alignment, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if args.semantic_map_skills:
        args.semantic_map_out.parent.mkdir(parents=True, exist_ok=True)
        args.semantic_map_out.write_text(
            json.dumps(semantic_audit, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(f"Module rows scored: {len(module_alignment)}")
    print(f"Group rows scored: {len(group_alignment)}")
    print(f"Programme rows scored: {len(programme_alignment)}")
    print(f"Job groups (SSOC-3): {len(job_profiles)}")
    print(f"Candidate groups for listing rerank: {args.candidate_groups_for_listings}")
    print(f"Top job listings per module: {args.top_job_listings}")
    print(f"Module input: {args.module_skills}")
    print(f"Module source type: {module_source_type}")
    print(f"Aux module skill sources used: {len(extra_skill_maps)}")
    print(f"Aux-support threshold for adding skills: {max(1, args.module_extra_skills_min_support)}")
    print(f"Modules matched in aux sources: {merge_stats['modules_with_extra_source_hit']}")
    print(f"Total base unique module-skills (pre-merge): {merge_stats['base_unique_skill_total']}")
    print(f"Total merged unique module-skills (post-merge): {merge_stats['merged_unique_skill_total']}")
    print(f"Freshgrad only: {args.freshgrad_only}")
    print(f"Job skill field: {args.job_skill_field}")
    print(f"Semantic skill mapping: {args.semantic_map_skills}")
    if args.semantic_map_skills:
        print(f"Semantic model: {args.semantic_model}")
        print(f"Semantic local-only: {args.semantic_local_only}")
        print(f"Semantic threshold: {args.semantic_threshold}")
        print(f"Semantic mapping rows: {len(semantic_audit)}")
        print(f"Saved semantic mapping -> {args.semantic_map_out}")
    print(f"Saved module alignment -> {args.module_alignment_out}")
    print(f"Saved group alignment -> {args.group_alignment_out}")
    print(f"Saved programme alignment -> {args.programme_alignment_out}")


if __name__ == "__main__":
    main()
