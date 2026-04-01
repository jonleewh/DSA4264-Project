import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_JOBS_DIR = PROJECT_ROOT / "data" / "data"
DEFAULT_SSOC_XLSX = PROJECT_ROOT / "data" / "ssoc2024-detailed-definitions.xlsx"
DEFAULT_OUTPUT_JSONL = PROJECT_ROOT / "data" / "test" / "job_ssoc345_with_skills_from_original.jsonl"
DEFAULT_OUTPUT_JSON = PROJECT_ROOT / "data" / "test" / "job_ssoc345_with_skills_from_original.json"


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_ssoc_title_lookup(xlsx_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = normalize_text(row[0] if len(row) > 0 else "")
        if not (code.isdigit() and len(code) in {3, 4, 5}):
            continue

        lookup[code] = normalize_text(row[1] if len(row) > 1 else "")
    return lookup


def parse_ssoc_code(raw) -> tuple[str | None, str | None, str | None]:
    text = normalize_text(raw)
    if not text:
        return None, None, None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 5:
        return None, None, None
    ssoc5 = digits[:5]
    ssoc4 = digits[:4]
    ssoc3 = digits[:3]
    return ssoc5, ssoc4, ssoc3


def extract_skill_names(skills_raw) -> list[str]:
    if not isinstance(skills_raw, list):
        return []
    names = []
    seen = set()
    for item in skills_raw:
        if isinstance(item, dict):
            skill = normalize_text(item.get("skill"))
        else:
            skill = normalize_text(item)
        if not skill:
            continue
        key = skill.lower()
        if key in seen:
            continue
        seen.add(key)
        names.append(skill)
    return names


def main():
    parser = argparse.ArgumentParser(
        description="Extract 3-digit SSOC from original job JSON files and map to SSOC 2024 Excel."
    )
    parser.add_argument("--jobs-dir", type=Path, default=DEFAULT_JOBS_DIR)
    parser.add_argument("--ssoc-xlsx", type=Path, default=DEFAULT_SSOC_XLSX)
    parser.add_argument("--output-jsonl", type=Path, default=DEFAULT_OUTPUT_JSONL)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    args = parser.parse_args()

    if not args.jobs_dir.exists():
        raise FileNotFoundError(f"Jobs directory not found: {args.jobs_dir}")
    if not args.ssoc_xlsx.exists():
        raise FileNotFoundError(f"SSOC Excel not found: {args.ssoc_xlsx}")

    ssoc_title_lookup = load_ssoc_title_lookup(args.ssoc_xlsx)
    if not ssoc_title_lookup:
        raise RuntimeError("No SSOC 3/4/5-digit rows found in the Excel file.")

    mapped_rows = []
    skipped_no_ssoc = 0
    skipped_no_match = 0

    files = sorted(args.jobs_dir.glob("*.json"))
    for i, path in enumerate(files, start=1):
        try:
            row = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue

        ssoc5, ssoc4, ssoc3 = parse_ssoc_code(row.get("ssocCode"))
        if not ssoc5:
            skipped_no_ssoc += 1
            continue

        title5 = ssoc_title_lookup.get(ssoc5, "")
        title4 = ssoc_title_lookup.get(ssoc4, "")
        title3 = ssoc_title_lookup.get(ssoc3, "")
        if not title3:
            skipped_no_match += 1
            continue

        metadata = row.get("metadata") or {}
        mapped_rows.append(
            {
                "job_post_id": metadata.get("jobPostId") or path.stem,
                "ssoc_5d_code": ssoc5,
                "ssoc_5d_title": title5,
                "ssoc_4d_code": ssoc4,
                "ssoc_4d_title": title4,
                "ssoc_3d_code": ssoc3,
                "ssoc_3d_title": title3,
                "skills": extract_skill_names(row.get("skills")),
            }
        )

        if i % 5000 == 0 or i == len(files):
            print(f"Scanned {i}/{len(files)} files...", flush=True)

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as f:
        for row in mapped_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(mapped_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Total files scanned: {len(files)}")
    print(f"Mapped rows: {len(mapped_rows)}")
    print(f"Skipped (missing ssocCode): {skipped_no_ssoc}")
    print(f"Skipped (3-digit hierarchy not in SSOC file): {skipped_no_match}")
    print(f"Saved JSONL: {args.output_jsonl}")
    print(f"Saved JSON: {args.output_json}")


if __name__ == "__main__":
    main()
