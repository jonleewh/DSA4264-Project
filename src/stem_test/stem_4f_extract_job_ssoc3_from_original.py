import argparse
import json
import pickle
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
FULL_DIR = PROJECT_ROOT / "data" / "stem_full"
DEFAULT_CLEANED_JOBS_INPUT = (
    PROJECT_ROOT / "data" / "cleaned_data" / "jobs_cleaned_portable.jsonl"
)
DEFAULT_SSOC_XLSX = PROJECT_ROOT / "data" / "ssoc2020.xlsx"
DEFAULT_OUTPUT_JSONL = FULL_DIR / "job_ssoc345_with_skills_from_original_STEM.jsonl"
DEFAULT_OUTPUT_JSON = FULL_DIR / "job_ssoc345_with_skills_from_original_STEM.json"


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_ssoc_title_lookup(xlsx_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = normalize_text(row[0] if len(row) > 0 else "")
        if not (code.isdigit() and len(code) in {3, 4, 5}):
            continue

        lookup[code] = normalize_text(row[1] if len(row) > 1 else "")
    return lookup


def parse_ssoc_code(raw) -> tuple[str | None, str | None, str | None]:
    text = normalize_text(raw)
    if not text:
        return None, None, None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 5:
        return None, None, None
    ssoc5 = digits[:5]
    ssoc4 = digits[:4]
    ssoc3 = digits[:3]
    return ssoc5, ssoc4, ssoc3


def extract_skill_names(row: dict) -> list[str]:
    names = []
    seen = set()
    for key in ("skills", "all_relevant_skills", "hard_skills", "soft_skills"):
        for item in row.get(key) or []:
            skill = normalize_text(item)
            if not skill:
                continue
            skill_key = skill.lower()
            if skill_key in seen:
                continue
            seen.add(skill_key)
            names.append(skill)
    return names


class _NumpyCompatUnpickler(pickle.Unpickler):
    """Allow loading pickles created under a different NumPy internal module path."""

    MODULE_MAP = {
        "numpy._core": "numpy.core",
        "numpy._core.numeric": "numpy.core.numeric",
        "numpy._core.multiarray": "numpy.core.multiarray",
    }

    def find_class(self, module: str, name: str) -> Any:
        module = self.MODULE_MAP.get(module, module)
        return super().find_class(module, name)


def _to_row_records(payload: Any, input_path: Path) -> list[dict]:
    if isinstance(payload, pd.DataFrame):
        return payload.to_dict("records")
    if isinstance(payload, list):
        return payload
    raise ValueError(
        f"Unsupported payload type loaded from {input_path}: {type(payload).__name__}. "
        "Expected pandas DataFrame or list[dict]."
    )


def _load_pickle_rows(input_path: Path) -> list[dict]:
    try:
        return _to_row_records(pd.read_pickle(input_path), input_path)
    except ModuleNotFoundError as exc:
        if "numpy._core" not in str(exc):
            raise
        with input_path.open("rb") as f:
            payload = _NumpyCompatUnpickler(f).load()
        return _to_row_records(payload, input_path)


def _try_fallback_jsonl_from_pickle(input_path: Path) -> list[dict] | None:
    jsonl_path = input_path.with_suffix(".jsonl")
    if jsonl_path.exists():
        print(
            "Falling back to JSONL input due to pickle compatibility issue: "
            f"{jsonl_path}",
            flush=True,
        )
        try:
            return load_cleaned_rows(jsonl_path)
        except Exception:
            return None
    return None


def load_cleaned_rows(input_path: Path) -> list[dict]:
    suffix = input_path.suffix.lower()
    if suffix == ".pkl":
        try:
            return _load_pickle_rows(input_path)
        except (ModuleNotFoundError, NotImplementedError, ValueError) as exc:
            fallback_rows = _try_fallback_jsonl_from_pickle(input_path)
            if fallback_rows is not None:
                return fallback_rows
            raise RuntimeError(
                "Failed to load pickle input and no sibling .jsonl fallback was found. "
                f"Input: {input_path}"
            ) from exc
    if suffix == ".json":
        payload = json.loads(input_path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(
                f"Expected a JSON array in {input_path}, got {type(payload).__name__}."
            )
        return payload
    if suffix == ".jsonl":
        with input_path.open("rb") as f:
            header = f.read(2)
        if header and header[:1] == b"\x80":
            raise ValueError(
                f"{input_path} has .jsonl extension but appears to be a pickle binary."
            )
        rows = []
        with input_path.open("r", encoding="utf-8") as f:
            for line_no, line in enumerate(f, start=1):
                raw = line.strip()
                if not raw:
                    continue
                try:
                    row = json.loads(raw)
                except json.JSONDecodeError as exc:
                    raise ValueError(
                        f"Invalid JSONL record at {input_path}:{line_no}: {exc.msg}"
                    ) from exc
                if not isinstance(row, dict):
                    raise ValueError(
                        f"Expected JSON object per JSONL line at {input_path}:{line_no}, "
                        f"got {type(row).__name__}."
                    )
                rows.append(row)
        return rows

    raise ValueError(
        f"Unsupported input format: {input_path}. Expected .pkl, .json, or .jsonl."
    )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Extract SSOC hierarchy from cleaned jobs and map to SSOC 2020 Excel, "
            "then generate SSOC STEM datasets."
        )
    )
    parser.add_argument("--cleaned-jobs", type=Path, default=DEFAULT_CLEANED_JOBS_INPUT)
    parser.add_argument("--ssoc-xlsx", type=Path, default=DEFAULT_SSOC_XLSX)
    parser.add_argument("--output-jsonl", type=Path, default=DEFAULT_OUTPUT_JSONL)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--include-non-fresh", action="store_true")
    args = parser.parse_args()

    if not args.cleaned_jobs.exists():
        raise FileNotFoundError(f"Cleaned jobs input not found: {args.cleaned_jobs}")
    if not args.ssoc_xlsx.exists():
        raise FileNotFoundError(f"SSOC Excel not found: {args.ssoc_xlsx}")

    ssoc_title_lookup = load_ssoc_title_lookup(args.ssoc_xlsx)
    if not ssoc_title_lookup:
        raise RuntimeError("No SSOC 3/4/5-digit rows found in the Excel file.")

    cleaned_rows = load_cleaned_rows(args.cleaned_jobs)

    fresh_only = not args.include_non_fresh
    mapped_rows: list[dict] = []
    skipped_no_ssoc = 0
    skipped_no_match = 0
    skipped_non_fresh = 0

    total = len(cleaned_rows)
    for i, row in enumerate(cleaned_rows, start=1):
        has_fresh_flag = "is_freshgrad" in row
        if fresh_only and has_fresh_flag and not row.get("is_freshgrad"):
            skipped_non_fresh += 1
            continue

        ssoc5, ssoc4, ssoc3 = parse_ssoc_code(row.get("ssoc_code") or row.get("ssocCode"))
        if not ssoc5:
            skipped_no_ssoc += 1
            continue

        title5 = ssoc_title_lookup.get(ssoc5, "")
        title4 = ssoc_title_lookup.get(ssoc4, "")
        title3 = ssoc_title_lookup.get(ssoc3, "")
        if not title3:
            skipped_no_match += 1
            continue

        mapped_rows.append(
            {
                "job_post_id": row.get("id") or row.get("uuid"),
                "ssoc_5d_code": ssoc5,
                "ssoc_5d_title": title5,
                "ssoc_4d_code": ssoc4,
                "ssoc_4d_title": title4,
                "ssoc_3d_code": ssoc3,
                "ssoc_3d_title": title3,
                "skills": extract_skill_names(row),
            }
        )

        if i % 5000 == 0 or i == total:
            print(f"Scanned {i}/{total} rows...", flush=True)

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as f:
        for mapped_row in mapped_rows:
            f.write(json.dumps(mapped_row, ensure_ascii=False) + "\n")

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(mapped_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Total rows scanned: {total}")
    print(f"Mapped rows: {len(mapped_rows)}")
    if fresh_only and any("is_freshgrad" in row for row in cleaned_rows):
        print(f"Skipped (non-fresh rows): {skipped_non_fresh}")
    print(f"Skipped (missing ssocCode): {skipped_no_ssoc}")
    print(f"Skipped (3-digit hierarchy not in SSOC file): {skipped_no_match}")
    print(f"Saved JSONL: {args.output_jsonl}")
    print(f"Saved JSON: {args.output_json}")


if __name__ == "__main__":
    main()
